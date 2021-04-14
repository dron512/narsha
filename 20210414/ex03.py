from openpyxl import Workbook
import random
wb = Workbook()
ws = wb.active

ws.append(["aa","bb","cc"])
ws.append(["dd","ee","ff"])
ws.append(["qq","hh","gg"])

for i in range(1,10):
    ws.append([random.randint(1,10),random.randint(1,30),random.randint(1,20)])

colB = ws["B"]
print(colB)
for i in colB:
    # print('i.value = '+str(i.value))
    print('i.value = ',i.value)
print("=================== REACT render(){ return (<div> </div>)}")
colBC = ws["B:C"]
for i in colBC:
    for j in i:
        print('j.value = '+str(j.value))

wb.save("b.xlsx")