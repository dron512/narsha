from openpyxl import load_workbook

wb = load_workbook("a.xlsx")
ws = wb.active

# for rw in range(1,11):
#     for col in range(1,11):
#         print(ws.cell(column=col,row=rw))
#         print(ws.cell(column=col,row=rw).value)

for rw in range(1,ws.max_row):
    for col in range(1,ws.max_column):
        print(ws.cell(column=col,row=rw))
        print(ws.cell(column=col,row=rw).value)

wb.close()