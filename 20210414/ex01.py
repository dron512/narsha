from openpyxl import Workbook # openpyxl 폴더에서 Workbook 클래스 가져오기
import random

# openpyxl 라이브러리 설치하기
# pip install openpyxl
wb = Workbook() # 임의의 엑셀 파일 생성
ws = wb.active  # 엑셀파일에서 첫번째 시트 가져오기
ws.title = "타이틀"

ws["A1"] = 1
ws["A2"] = 2

ws.cell(column=2,row=1).value = 10
ws.cell(column=2,row=2).value = 20

for i in range(1,11):    # 11미만까지이기 때문에 10까지... RND 0<X<1
    for j in range(1,11):
        # ws.cell(column=i,row=j).value = random.randint(1,100)
        ws.cell(column=i,row=j, value = random.randint(1,100))

wb.save("a.xlsx") # 파일 열었으면 다른 사람들 이용할 수 있게 저장하고 닫기