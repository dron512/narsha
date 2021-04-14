from openpyxl import Workbook

wb = Workbook() # 엑셀 열기
ws = wb.active  # 첫번째 시트 가져와서 ws 변수넣기

ws.title = "시트제목"
ws['A1'] = '시트내용'
ws['A2'] = 'A2'

print(ws['A1'])
print(ws['A2'])

print(ws['A1'].value)
print(ws['A2'].value)

wb.save('a.xlsx')

wb.close()